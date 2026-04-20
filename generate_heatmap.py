"""
Chat Heatmap Generator
======================
Drop your raw Yellow.ai / chat export CSV here and run:

    python generate_heatmap.py your_export.csv

Outputs:
  - chat_heatmap_output.xlsx   (Excel with Config + Heatmap + Shrinkage sheets)
  - index.html                 (Web dashboard — host on GitHub Pages)

What this does automatically:
  1. Parses MM/DD/YYYY hh:mm:ss AM/PM timestamps — no manual fixing needed
  2. Derives Date, Day of Week, Hour, 30-min Slot from each ticket
  3. Builds normalised pivot (avg chats per slot per day)
  4. Calculates Agents Needed using your AHT and Concurrency
  5. Writes formatted Excel + interactive HTML dashboard

Edit the CONFIG section below to change AHT, concurrency, shifts, etc.
"""

import sys
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════
AHT_MINUTES   = 16
CONCURRENCY   = 3.0
SLOT_MINUTES  = 30
ROSTERED      = 39
UNPLN_PCT     = 0.10

SHIFTS = [
    ("S1  7AM-4PM",   7,  16, 13),
    ("S2  3PM-12AM", 15,  24, 10),
    ("S3  6PM-3AM",  18,  27, 10),
    ("S4 11PM-8AM",  23,  32,  6),
]

OUTPUT_EXCEL = "chat_heatmap_output.xlsx"
OUTPUT_HTML  = "index.html"

DAYS_ORDER = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
DAYS_SHORT = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]

# ══════════════════════════════════════════════════════════════════════════════
# PARSE
# ══════════════════════════════════════════════════════════════════════════════
def load_and_parse(filepath):
    print(f"Loading: {filepath}")
    df = pd.read_csv(filepath, low_memory=False, encoding="utf-8-sig")
    print(f"  {len(df):,} rows, {len(df.columns)} columns")

    ts_candidates = ["initialized_time", "assigned_time", "opened_time", "Created time"]
    ts_col = next((c for c in ts_candidates if c in df.columns), None)
    if not ts_col:
        raise ValueError(f"No timestamp column found. Expected one of: {ts_candidates}")
    print(f"  Using timestamp column: '{ts_col}'")

    def parse_ts(val):
        if pd.isna(val) or str(val).strip() in ("", "N/A", "null"):
            return pd.NaT
        for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S",
                    "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(str(val).strip(), fmt)
            except ValueError:
                continue
        return pd.NaT

    print("  Parsing timestamps...")
    df["_ts"] = df[ts_col].apply(parse_ts)
    bad = df["_ts"].isna().sum()
    if bad:
        print(f"  Warning: {bad:,} rows skipped (unparseable timestamps)")
    df = df.dropna(subset=["_ts"]).copy()
    print(f"  {len(df):,} rows with valid timestamps")

    df["_date"] = df["_ts"].dt.date
    df["_dow"]  = df["_ts"].dt.day_name()
    df["_slot"] = df["_ts"].apply(
        lambda x: f"{x.hour:02d}:{(x.minute // SLOT_MINUTES) * SLOT_MINUTES:02d}"
    )
    return df


# ══════════════════════════════════════════════════════════════════════════════
# PIVOT
# ══════════════════════════════════════════════════════════════════════════════
def build_pivot(df):
    slots_per_day = (24 * 60) // SLOT_MINUTES
    all_slots = [
        f"{(i * SLOT_MINUTES) // 60:02d}:{(i * SLOT_MINUTES) % 60:02d}"
        for i in range(slots_per_day)
    ]
    counts = df.groupby(["_slot", "_dow"]).size().unstack(fill_value=0)
    counts = counts.reindex(columns=DAYS_ORDER, fill_value=0)
    counts = counts.reindex(all_slots, fill_value=0)

    day_counts = df.groupby("_dow")["_date"].nunique()
    for day in DAYS_ORDER:
        n = max(day_counts.get(day, 1), 1)
        counts[day] = (counts[day] / n).round(0).astype(int)

    counts["Daily Avg"]     = counts[DAYS_ORDER].mean(axis=1).round(0).astype(int)
    counts["Agents Needed"] = (counts["Daily Avg"] * AHT_MINUTES / SLOT_MINUTES / CONCURRENCY).round(1)

    date_range = f"{df['_date'].min()} to {df['_date'].max()}"
    print(f"  Date range: {date_range}")
    print(f"  Peak avg: {counts['Daily Avg'].max()} chats/slot")
    return counts, all_slots, date_range


# ══════════════════════════════════════════════════════════════════════════════
# SHIFT HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def agents_online(slot_str):
    h, m = map(int, slot_str.split(":"))
    mins = h * 60 + m
    total = 0
    for _, s_h, e_h, cnt in SHIFTS:
        s, e = s_h * 60, e_h * 60
        if s <= mins < e or s <= (mins + 24 * 60) < e:
            total += cnt
    return total

def shift_label(slot_str):
    h, m = map(int, slot_str.split(":"))
    mins = h * 60 + m
    active = []
    for name, s_h, e_h, _ in SHIFTS:
        s, e = s_h * 60, e_h * 60
        if s <= mins < e or s <= (mins + 24 * 60) < e:
            active.append(name.split()[0])
    return "+".join(active) if active else "--"


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def bdr():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def xcell(ws, row, col, val, bg=None, fg="000000", bold=False, size=9,
          align="center", italic=False, wrap=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fg, size=size, name="Arial", italic=italic)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = bdr()
    if num_fmt:
        c.number_format = num_fmt
    return c

def xhdr(ws, row, col, val, bg="2E75B6", fg="FFFFFF", size=9):
    return xcell(ws, row, col, val, bg=bg, fg=fg, bold=True, size=size, wrap=True)

def heat_color(val, vmin, vmax):
    if vmax == vmin:
        return "FFFFFF"
    ratio = (val - vmin) / (vmax - vmin)
    if ratio < 0.2:
        r = int(210 + ratio*5*45); g = int(228 + ratio*5*10); b = 255
    elif ratio < 0.4:
        r = 255; g = int(238-(ratio-0.2)*5*40); b = int(255-(ratio-0.2)*5*80)
    elif ratio < 0.6:
        r = 255; g = int(198-(ratio-0.4)*5*50); b = int(175-(ratio-0.4)*5*100)
    elif ratio < 0.8:
        r = 255; g = int(148-(ratio-0.6)*5*80); b = int(75-(ratio-0.6)*5*75)
    else:
        r = int(255-(ratio-0.8)*5*55); g = int(68-(ratio-0.8)*5*68); b = 0
    return f"{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}"


def build_config_sheet(wb):
    ws = wb.create_sheet("Config")
    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value="Configuration Panel  -  Edit yellow cells to recalculate the entire workbook")
    t.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t.fill = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    n = ws.cell(row=2, column=1, value="Change the yellow cells. All heatmap and shrinkage calculations update automatically.")
    n.font = Font(italic=True, size=9, color="444444", name="Arial")
    n.fill = PatternFill("solid", fgColor="EBF3FB")
    n.alignment = Alignment(horizontal="center", vertical="center")

    for ci, h in enumerate(["Named Cell","Parameter","Value","Unit","Notes"]):
        xhdr(ws, 5, ci+1, h)
    for ci, w in enumerate([18,30,14,12,52]):
        ws.column_dimensions[get_column_letter(ci+1)].width = w

    params = [
        ("AHT",       "AHT (Average Handle Time)", AHT_MINUTES,  "minutes", "Average time per chat in minutes"),
        ("CONC",      "Concurrency",                CONCURRENCY,  "chats",   "Simultaneous chats per agent. Typical: 2.5-4.0"),
        ("SLOT_MINS", "Slot Size",                  SLOT_MINUTES, "minutes", "Heatmap granularity (30 min recommended)"),
        ("ROSTER",    "Rostered Headcount",         ROSTERED,     "agents",  "Total agents on roster"),
        ("UNPLN_PCT", "Unplanned Shrinkage %",      UNPLN_PCT,    "",        "Sick calls / no-shows. Benchmark: 8-12%"),
    ]
    for ri, (name, label, default, unit, note) in enumerate(params):
        r = ri + 6
        ws.row_dimensions[r].height = 22
        xcell(ws, r, 1, name, bg="F2F2F2", bold=True, size=9, fg="1F4E79", align="left")
        xcell(ws, r, 2, label, align="left", size=9)
        vc = ws.cell(row=r, column=3, value=default)
        vc.font = Font(bold=True, size=11, name="Arial", color="1F4E79")
        vc.fill = PatternFill("solid", fgColor="FFFF00")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = bdr()
        if name == "UNPLN_PCT":
            vc.number_format = "0%"
        xcell(ws, r, 4, unit, size=9, fg="595959", italic=True)
        xcell(ws, r, 5, note, align="left", size=8, italic=True, fg="595959")

    ws.merge_cells("A12:E12")
    dh = ws.cell(row=12, column=1, value="DERIVED VALUES  (auto-calculated - do not edit)")
    dh.font = Font(bold=True, size=10, color="FFFFFF", name="Arial")
    dh.fill = PatternFill("solid", fgColor="375623")
    dh.alignment = Alignment(horizontal="left", vertical="center")

    for ri, (label, formula, fmt) in enumerate([
        ("Effective Roster", "=Config!C9*(1-Config!C10)", "0.0"),
        ("Shrinkage Ratio",  "=1-Config!C10",             "0%"),
        ("Formula note",     "= (Avg Chats x AHT) / Slot_mins / Concurrency", None),
    ]):
        r = ri + 13
        xcell(ws, r, 1, label, align="left", size=9, bold=True, bg="EBF5FB")
        vc = ws.cell(row=r, column=3, value=formula)
        is_formula = formula.startswith("=")
        vc.font = Font(bold=is_formula, italic=not is_formula,
                       size=10 if is_formula else 8, name="Arial",
                       color="375623" if is_formula else "595959")
        vc.fill = PatternFill("solid", fgColor="E2EFDA")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = bdr()
        if fmt and is_formula:
            vc.number_format = fmt

    ws.merge_cells("A17:E17")
    tip = ws.cell(row=17, column=1,
        value="TIP: Change AHT or Concurrency above, press Enter, and Agents Needed updates across all sheets")
    tip.font = Font(italic=True, size=9, name="Arial", color="1F4E79")
    tip.fill = PatternFill("solid", fgColor="EBF3FB")
    tip.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A6"


def build_heatmap_sheet(wb, pivot, all_slots, date_range):
    ws = wb.create_sheet("Half-Hour Heatmap")
    AHT_REF  = "Config!C6"
    CONC_REF = "Config!C7"
    SLOT_REF = "Config!C8"

    ws.merge_cells(f"A1:{get_column_letter(14)}1")
    t = ws.cell(row=1, column=1,
        value=f"Chat Volume Heatmap  |  {SLOT_MINUTES}-Min Slots  |  {date_range}  |  Edit AHT & Concurrency in Config sheet")
    t.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    t.fill = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Slot"] + DAYS_SHORT + ["Daily\nAvg","Agents\nNeeded","Online\n(Shift)","Gap\n(+/-)","Status","Shift\nWindow"]
    widths   = [7] + [8]*7 + [9,10,8,7,17,13]
    for ci, (h, w) in enumerate(zip(headers, widths)):
        bg = "833C00" if ci == 9 else "2E75B6"
        xhdr(ws, 3, ci+1, h, bg=bg)
        ws.column_dimensions[get_column_letter(ci+1)].width = w
    ws.row_dimensions[3].height = 30

    all_vals = [int(pivot.loc[s, d]) for s in all_slots for d in DAYS_ORDER if s in pivot.index]
    vmin, vmax = min(all_vals), max(all_vals)

    for ri, slot in enumerate(all_slots):
        r = ri + 4
        ws.row_dimensions[r].height = 16
        xcell(ws, r, 1, slot, bold=True, size=9)

        for ci, day in enumerate(DAYS_ORDER):
            val = int(pivot.loc[slot, day]) if slot in pivot.index else 0
            bg = heat_color(val, vmin, vmax)
            ratio = (val - vmin) / (vmax - vmin) if vmax > vmin else 0
            c = ws.cell(row=r, column=ci+2, value=val)
            c.font = Font(size=8, name="Arial", color="FFFFFF" if ratio > 0.72 else "000000")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr()

        avg = int(pivot.loc[slot, "Daily Avg"]) if slot in pivot.index else 0
        avg_ratio = (avg - vmin) / (vmax - vmin) if vmax > vmin else 0
        ac = ws.cell(row=r, column=9, value=avg)
        ac.font = Font(bold=True, size=9, name="Arial", color="FFFFFF" if avg_ratio > 0.72 else "000000")
        ac.fill = PatternFill("solid", fgColor=heat_color(avg, vmin, vmax))
        ac.alignment = Alignment(horizontal="center", vertical="center")
        ac.border = bdr()

        nc = ws.cell(row=r, column=10, value=f"=ROUND(I{r}*{AHT_REF}/{SLOT_REF}/{CONC_REF},1)")
        nc.font = Font(bold=True, size=9, name="Arial", color="833C00")
        nc.fill = PatternFill("solid", fgColor="FCE4D6")
        nc.alignment = Alignment(horizontal="center", vertical="center")
        nc.border = bdr(); nc.number_format = "0.0"

        online = agents_online(slot)
        oc = ws.cell(row=r, column=11, value=online)
        oc.font = Font(bold=True, size=9, name="Arial")
        oc.alignment = Alignment(horizontal="center", vertical="center")
        oc.border = bdr()

        gc = ws.cell(row=r, column=12, value=f"=ROUND(K{r}-J{r},1)")
        gc.font = Font(bold=True, size=9, name="Arial")
        gc.alignment = Alignment(horizontal="center", vertical="center")
        gc.border = bdr(); gc.number_format = "0.0"

        sf = (f'=IF(L{r}>=3,"Comfortable",'
              f'IF(L{r}>=1,"Good",'
              f'IF(L{r}>=0,"Tight",'
              f'IF(L{r}>=-2,"Under","Critical"))))')
        sc2 = ws.cell(row=r, column=13, value=sf)
        sc2.font = Font(size=8, name="Arial")
        sc2.alignment = Alignment(horizontal="center", vertical="center")
        sc2.border = bdr()

        xcell(ws, r, 14, shift_label(slot), size=7, italic=True, bg="F0F0F0")

    ws.freeze_panes = "B4"


def build_shrinkage_sheet(wb, pivot, all_slots):
    ws = wb.create_sheet("Shrinkage Planner")
    AHT_REF   = "Config!C6"
    CONC_REF  = "Config!C7"
    SLOT_REF  = "Config!C8"
    UNPLN_REF = "Config!C10"

    ws.merge_cells("A1:K1")
    t = ws.cell(row=1, column=1,
        value="Shrinkage Planner  |  Edit AHT, Concurrency & Shrinkage in Config sheet  |  Auto-recalculates")
    t.font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
    t.fill = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")

    cols = ["Slot","Avg\nChats","Agents\nNeeded","Gross\nOnline","Unplanned\nLoss",
            "Net After\nUnplanned","Planned\nOff","Effective\nHC","Gap","Break\nPolicy","Status"]
    widths = [7,9,11,9,10,11,9,10,7,20,16]
    for ci, (h, w) in enumerate(zip(cols, widths)):
        xhdr(ws, 3, ci+1, h)
        ws.column_dimensions[get_column_letter(ci+1)].width = w
    ws.row_dimensions[3].height = 36

    all_avg = [int(pivot.loc[s, "Daily Avg"]) for s in all_slots if s in pivot.index]
    vmin_a, vmax_a = min(all_avg), max(all_avg)

    for ri, slot in enumerate(all_slots):
        r = ri + 4
        ws.row_dimensions[r].height = 17

        avg = int(pivot.loc[slot, "Daily Avg"]) if slot in pivot.index else 0
        ratio = (avg - vmin_a) / (vmax_a - vmin_a) if vmax_a > vmin_a else 0
        if ratio < 0.25:
            poff, brec, bbg = 2, "Safe - 2 agents on break", "D5F5E3"
        elif ratio < 0.55:
            poff, brec, bbg = 1, "Max 1 agent on break",     "FEF9E7"
        else:
            poff, brec, bbg = 0, "Block - peak slot",        "FADBD8"

        online = agents_online(slot)

        xcell(ws, r, 1, slot, bold=True, size=9)
        xcell(ws, r, 2, avg)

        c3 = ws.cell(row=r, column=3, value=f"=ROUND(B{r}*{AHT_REF}/{SLOT_REF}/{CONC_REF},1)")
        c3.font = Font(bold=True, size=9, name="Arial", color="833C00")
        c3.fill = PatternFill("solid", fgColor="FCE4D6")
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c3.border = bdr(); c3.number_format = "0.0"

        xcell(ws, r, 4, online, bg="EBF5FB")

        c5 = ws.cell(row=r, column=5, value=f'="-"&TEXT(ROUND(D{r}*{UNPLN_REF},1),"0.0")')
        c5.font = Font(bold=True, size=9, name="Arial", color="922B21")
        c5.fill = PatternFill("solid", fgColor="FADBD8")
        c5.alignment = Alignment(horizontal="center", vertical="center")
        c5.border = bdr()

        c6 = ws.cell(row=r, column=6, value=f"=ROUND(D{r}*(1-{UNPLN_REF}),1)")
        c6.font = Font(bold=True, size=9, name="Arial", color="1A5276")
        c6.fill = PatternFill("solid", fgColor="EBF5FB")
        c6.alignment = Alignment(horizontal="center", vertical="center")
        c6.border = bdr(); c6.number_format = "0.0"

        xcell(ws, r, 7, f"-{poff}" if poff>0 else "0",
              bg="D5F5E3" if poff==2 else ("FEF9E7" if poff==1 else "FADBD8"),
              bold=True, fg="1A5276" if poff>0 else "7B241C")

        c8 = ws.cell(row=r, column=8, value=f"=ROUND(F{r}-G{r},1)")
        c8.font = Font(bold=True, size=9, name="Arial")
        c8.alignment = Alignment(horizontal="center", vertical="center")
        c8.border = bdr(); c8.number_format = "0.0"

        c9 = ws.cell(row=r, column=9, value=f"=ROUND(H{r}-C{r},1)")
        c9.font = Font(bold=True, size=9, name="Arial")
        c9.alignment = Alignment(horizontal="center", vertical="center")
        c9.border = bdr(); c9.number_format = "0.0"

        xcell(ws, r, 10, brec, bg=bbg, size=8, italic=True)

        sf = (f'=IF(I{r}>=4,"Comfortable",'
              f'IF(I{r}>=1,"Good",'
              f'IF(I{r}>=0,"Tight",'
              f'IF(I{r}>=-2,"Under","Critical"))))')
        c11 = ws.cell(row=r, column=11, value=sf)
        c11.font = Font(size=8, name="Arial")
        c11.alignment = Alignment(horizontal="center", vertical="center")
        c11.border = bdr()

    ws.freeze_panes = "B4"


def build_excel(pivot, all_slots, date_range):
    wb = Workbook()
    wb.remove(wb.active)
    build_config_sheet(wb)
    build_heatmap_sheet(wb, pivot, all_slots, date_range)
    build_shrinkage_sheet(wb, pivot, all_slots)
    wb.save(OUTPUT_EXCEL)
    print(f"  Excel saved: {OUTPUT_EXCEL}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    if len(sys.argv) < 2:
        print("Usage:  python generate_heatmap.py <your_chat_export.csv>")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    print("\n-- Chat Heatmap Generator --")
    print(f"AHT: {AHT_MINUTES}min  |  Concurrency: {CONCURRENCY}  |  Slot: {SLOT_MINUTES}min  |  Roster: {ROSTERED}\n")

    df = load_and_parse(filepath)
    print("\nBuilding pivot...")
    pivot, all_slots, date_range = build_pivot(df)

    print("\nWriting Excel...")
    build_excel(pivot, all_slots, date_range)

    print("Writing HTML dashboard...")
    from build_html import build_html
    build_html(pivot, all_slots, date_range, len(df),
               AHT_MINUTES, CONCURRENCY, SLOT_MINUTES, ROSTERED, SHIFTS,
               OUTPUT_HTML)

    print(f"\nDone!")
    print(f"  Excel:  {OUTPUT_EXCEL}")
    print(f"  Web:    {OUTPUT_HTML}")


if __name__ == "__main__":
    main()
